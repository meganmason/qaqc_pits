'''
This script pulls the coordinates (UTME, UTMN) from the SnowEx pit forms and
writes the info with header info to a csv.

Intended for quick extraction of UTMs as a check on conversion and transcriptions.

This script works for the __2020__SnowEx__Pit__Form__

This script was modified from the pits_xls2csv_v6.py script from C. Vuyovich.

'''

__author__ = "Megan Mason, ATAAerospace Data Management Liaison"
__version__ = "01"
__maintainer__ = "Megan Mason"
__email__ = "meganmason491@boisestate.edu"
__status__ = "Dvp"
__date__ = "01.2021"

import datetime
import glob
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from pyproj import Proj, transform
from csv import writer
import textwrap
import utm

#----------------------------------METHODS--------------------------------------
def readSnowpit(filename):

        # open excel file
        wb = load_workbook(filename)
        ws = wb.active

        # date / location / stie / pit-ID / coordinates
        if type(ws['S6'].value) is str:
            date = datetime.datetime.strptime(ws['S6'].value, '%m/%d/%Y') #remove this code block when ER and NW are in the correct datetime format
        else:
            date = (ws['S6'].value).date()

        location = ws['B2'].value
        site = ws['B4'].value
        # pitID = (ws['B6'].value).split('_')[0] # use when ER and NW are in correct format
        pitID = str(ws['B6'].value)[:6]

        lat = ws['L4'].value # easting
        lon = ws['Q4'].value # northing
        zone = ws['X4'].value


    # lat/lon corrections-------------------------------------------------------

        # empty??
        if lat == None:
            LAT = np.nan
            LON = np.nan

        else:

            # remove strings
            if type(lat) is str:
                lat = lat.split('.')[0] # cuts off anything after a decimal (otherwise it would join together and incr. mag. of num)
                lat = float(''.join([i for i in list(lat) if i.isdigit()]))
            if type(lon) is str:
                lon = lon.split('.')[0]
                lon = float(''.join([i for i in list(lon) if i.isdigit()]))

            # If lat = negative, swap lat/lon
            if lat < 0: #recorded as longitude (negative value)
                lat, lon = lon, lat # swap coards

            # convert UTMs to lat/lon
            if lat > 90: #recorded as UTMs
                UTME = lat
                UTMN = lon

                lat = utm.to_latlon(UTME, UTMN, zone, "Northern")[0] #tuple output, save first

            if lon > 0:
                lon = utm.to_latlon(UTME, UTMN, zone, "Northern")[1] #tuple output, save second


            LAT = round(lat, 5)
            LON = round(lon, 5)


        return date, location, site, pitID, LAT, LON  #lat, lon = NO corrections, #LAT, LON = with corrections
        # return LAT, LON
#----------------------------------BODY--------------------------------------

if __name__ == "__main__":

    # set-up
    path_in = Path('/Users/meganmason491/Google Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT')

    # data
    data = []
    column_lst = ['Date', 'Location', 'Site', 'PitID', 'latitude', 'longitude']
    # column_lst = ['latitude', 'longitude']

    for filename in path_in.rglob('*.xlsx'):

        # filename
        print(f"...reading {filename.name}")

        # append data list with each pit
        data.append(readSnowpit(filename))

    # convert data list to dataframe
    df = pd.DataFrame(data, columns=column_lst)

    print('No. of data rows:', len(df.index))

    # create csv
    df.to_csv('/Users/meganmason491/Documents/snowex/2020/timeseries/qaqc_pits/all_pits_raw_coords.csv', sep=',', header=True)
