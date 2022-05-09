import pandas as pd
import glob
from datetime import datetime
from csv import writer
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font

'''
This script formats the pit sheets given to us by Mark Raleigh's group.
These pits were cleaned by Kate Hale (UC Boulder).
    Cleaning steps included:
    - transfering Mark's pit sheets to the SnowEx template
    - gap filling lot of header data once confirmed with Mark and his student Hannah B.
    - adjusting some times to create a unique snow pit ID
    - putting grain size in the strat comments because grain bins didn't match
    - universonal changes, transcription errors, etc. caught by Kate.

Kate did a lot of copy/paste from the variety of pit sheets submitted by Mark's group.
Thus, the SnowEx template was overwritten. This script formats (mostly center aligns)
much of the pit sheet and a few other items listed here:
    - aligns text in cell
    - bolds neccesary text
    - standarizes font size, style
    - renames a few cells (Location, pitID, LWC serial number)
'''
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#static variables
#~~~ East River pits ~~~~
# Location = "East River" #Niwot
# loc_code = "ER" #NW
# site_code = "UP" # aspen, O2, upper, etc.

# ~~~ Niwot pits ~~~~~
Location = "Niwot" #Niwot
Site = 'Open Flat'
loc_code = "NW" #NW
site_code = "OF" # aspen, O2, upper, etc.

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/East-River-formatted')
path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/Niwot-formatted')
# for filename in path.rglob('*.xlsx'):
for filename in path.rglob('*OF??.xlsx*'): #('*Trench13*'):
    print(filename.name)

    # load workbook
    wb = load_workbook(filename)
    ws = wb.active

    # ws['J9'].alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)

# Format Header
    # alignment
    ws['B2'].alignment = Alignment(vertical='center', horizontal='left') # location
    ws['B4'].alignment = Alignment(vertical='center', horizontal='left')  # site
    ws['B6'].alignment = Alignment(vertical='center', horizontal='left')   # pitID
    ws['G6'].alignment = Alignment(vertical='center', horizontal='center') # HS
    ws['L6'].alignment = Alignment(vertical='center', horizontal='center') # Air Temp
    ws['N6'].alignment = Alignment(vertical='center', horizontal='center') # slope
    ws['Q6'].alignment = Alignment(vertical='center', horizontal='center') # aspect
    ws['S6'].alignment = Alignment(vertical='center', horizontal='center') # Date
    ws['L4'].alignment = Alignment(vertical='center', horizontal='center') # UTME
    ws['Q4'].alignment = Alignment(vertical='center', horizontal='center') # UTMN
    ws['X4'].alignment = Alignment(vertical='center', horizontal='center') # Zone
    ws['X6'].alignment = Alignment(vertical='center', horizontal='center') # Time
    ws['Y2'].alignment = Alignment(vertical='center', horizontal='center') # Comments

    # font size, name, bold
    font_size = Font(size=11, name='Arial', bold=True)
    ws['B2'].font = font_size
    ws['B4'].font = font_size
    ws['B6'].font = font_size
    ws['G6'].font = font_size
    ws['L6'].font = font_size
    ws['N6'].font = font_size
    ws['Q6'].font = font_size
    ws['S6'].font = font_size
    ws['L4'].font = font_size
    ws['Q4'].font = font_size
    ws['X4'].font = font_size
    ws['X6'].font = font_size
    ws['Y2'].font = font_size

    # format date (mm/dd/yyyy to yyyy-mm-dd and make datetime obj.)
    if type(ws['S6'].value) is str:
        dt = datetime.strptime(ws['S6'].value, "%m/%d/%Y") #"%Y-%m-%d"
        ws['S6'].value = dt.date()

    # rename Location
    ws['B2'].value = Location

    # rename Site
    if Location is "Niwot":
        ws['B4'].value = Site

    # modify pit ID (yyyymmdd)
    d_string = datetime.strftime(ws['S6'].value, "%Y%m%d")
    ws['B6'].value = 'CO' + loc_code + site_code + '_' + d_string

    # rename LWC serial number cell
    ws['I5'].value = 'WISe Serial Number'
    ws['I5'].alignment = Alignment(vertical='center', horizontal='center')
    ws['I5'].font = Font(size=11, name='Arial', bold=True)




# Core pit measurements

    # iterate over all rows for Density, LWC, Temp, and Strat
    for i in range(9,32): # i.e row 10-33 in excel workbook.

         # iterate over all columns
         for j in range(1,26): # i.e. columns ~B-X
              # get particular cell value
              cell_obj=ws.cell(row=i, column=j)
              # align
              cell_obj.alignment = Alignment(vertical='center', horizontal='center')

    for i in range(9,32): # i.e row 10-33 in excel workbook.
          # get particular cell value
          cell_obj=ws.cell(row=i, column=12) #Temp
          # bold temp interval
          cell_obj.font = Font(bold=True)

    for i in range(9,32): # i.e row 10-33 in excel workbook.
          # get particular cell value
          cell_obj=ws.cell(row=i, column=26) #Strat Comments
          # bold temp interval
          cell_obj.alignment = Alignment(vertical='center', horizontal='left')




# Save workbook
    wb.save(filename)


# ITERATE OVER ALL CELLS
#     # iterate over all cells
# # iterate over all rows
# for i in range(1,max_row+1):
#
#      # iterate over all columns
#      for j in range(1,max_column+1):
#           # get particular cell value
#           cell_obj=sheet.cell(row=i,column=j)
#           # print cell value
#           print(cell_obj.value,end=' | ')
#      # print new line
#      print('\n')
# link: https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
