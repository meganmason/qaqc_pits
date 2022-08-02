import datetime
import glob
import os
import shutil
import numpy as np
import pandas as pd
from csv import writer
from pathlib import Path
import utm
from openpyxl import load_workbook


path_in = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/interval-boards-only-sheets/')
# path_in = Path ('/Users/meganmason491/Downloads/test/')

def addTime(filename):

    # get time
    wb = load_workbook(filename)
    ws = wb.active
    timeString = ws['X6'].value.strftime('%H%M')

    # change filename
    parts = filename.name.split('_') #old fname split into parts
    parts.insert(2, timeString) #insert timeString based on 3rd position
    new_name = '_'.join(parts) # join the parts together

    return new_name

    # Same below, but for Bogus Lower pits when prepping interval board stuff.

    # # get time
    # wb = load_workbook(filename)
    # ws = wb.active
    # timeString = ws['X6'].value.strftime('%H%M')
    #
    # # change filename
    # parts = filename.stem.split('_') #old fname split into parts
    # parts.insert(2, timeString) #insert timeString based on 3rd position
    # new_name = '_'.join(parts) + filename.suffix # join the parts together
    #
    # return new_name


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
if __name__ == "__main__":

    for path in path_in.rglob('*.xlsx'):

        if 'formated' in path.name:
            pass

        else:
            if path.is_file():

                print('...processing: ', path.stem)

                # old name
                old_name = path.stem

                # original file extension
                directory = path.parent

                # add time to name
                new_name = addTime(path)

                # print("%s --> %s" %(old_name, new_name))

                # rename file
                path.rename(Path(directory, new_name))


#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
