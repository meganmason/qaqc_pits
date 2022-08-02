import pandas as pd
import glob
import os
import datetime
from csv import writer
from pathlib import Path
from openpyxl import load_workbook

'''
This script is to make sure all times are actually datetime.time types.

I thought this already happened, but having issues still! So here goes another attempt.

This catches:
string --> datetime.time
datetime.datetime --> datetime.time
datetime.time --> 'datetime.time' (...and for good measure)
'''

path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/COMPLETE/')

# daylight_savings = datetime(2020, 3, 8) # March 8th, 2020


for filename in sorted(path.rglob('*.xlsx')):

    # load workbook
    wb = load_workbook(filename)
    ws = wb.active

    # time variable
    time = ws['X6'].value # could be string, datetime.datetime or datetime.time type


    if isinstance(time, datetime.time): # really unneccessary since we want datetime.time
        f_time = time
        ws['X6'].value = f_time
        ws['X6'].number_format = 'HH:MM'
        print('type: {} for {} --> {}'.format(type(time), filename.stem, type(f_time)))

    elif isinstance(time, datetime.datetime):
        f_time = time.time()
        ws['X6'].value = f_time
        ws['X6'].number_format = 'HH:MM'
        print('type: {} for {} --> {}'.format(type(time), filename.stem, type(f_time)))

    elif isinstance(time, str):
        f_time = datetime.datetime.strptime(time, '%H:%M').time() # = datetime.time()
        ws['X6'].value = f_time
        ws['X6'].number_format = 'HH:MM'
        print('type: {} for {} --> {}'.format(type(time), filename.stem, type(f_time)))

    else:
        print('Hold up, DataType not accounted for\nfile is: {} and time datatype is {}'.format(filename.stem, type(time)))


    # Save workbook
    wb.save(filename)

print('all done')





# ~~~~~~~~~~~~~~~~~~~~might need later:
#if:
# pitID, date_str, time_str, desc = filename.stem.split('_') # split filename to isolate time_str
# newFilename = pitID + '_' + date_str + '_' + newTime.time().strftime("%H%M") + '_' + desc + filename.suffix # remove old time string, and replace with newTime
# newFilename = Path(pitID + '_' + date_str + '_' + newTime.time().strftime("%H%M") + '_' + desc + filename.suffix) # remove old time string, and replace with newTime
#
#else:
# newFilename = filename
