import pandas as pd
import glob
from datetime import datetime
from csv import writer
from pathlib import Path
from openpyxl import load_workbook

'''
This script is for the additional East River and Niwot snow pits.

This script removes grain type in the top row if it is 'PP'. This was decided because
the stratigraphy comments typically say 'fresh snow', 'powder', etc. and that is not
definitivily 'PP', we know it's possible to be 'DF' or even 'RG'.
'''

Location = "Niwot" #East-River

path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/{}-formatted'.format(Location))

for filename in path.rglob('*.xlsx'):
    print(filename.name)

    # load workbook
    wb = load_workbook(filename)
    ws = wb.active

    if ws['W10'].value == 'PP':
        ws['W10'].value = " "


# Save workbook
    wb.save(filename)
