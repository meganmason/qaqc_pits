import datetime
import glob
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from csv import writer
import textwrap

'''
This script cleans up the SnowEx 2020 Time Series Campaign Names:
    * location
    * site

            location = ws['B2'].value
            site = ws['B4'].value
            pitID = ws['B6'].value
'''

#----------------------------------BODY--------------------------------------

# paths
path_in = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/COMPLETE')

for filename in sorted(path_in.rglob('*.xlsx')):

    print(filename.name)

    # open excel file
    wb = load_workbook(filename)
    ws = wb.active

    # Location updates

    if ws['B6'].value[:4] == 'CAAM': # American --> American River Basin
        ws['B2'].value = 'American River Basin'

    if ws['B6'].value[:4] == 'COFE':  # Fraser --> Fraser Experimental Forest
        ws['B2'].value = 'Fraser Experimental Forest'

    if ws['B6'].value[:4] == 'UTLC':  # Little Cottonwood --> Little Cottonwood Canyon
        ws['B2'].value = 'Little Cottonwood Canyon'

    if ws['B6'].value[:4] == 'CONW': # Niwot --> Niwot Ridge
        ws['B2'].value = 'Niwot Ridge'

    if ws['B6'].value[:4] == 'CASH': # Sagehen --> Sagehen Creek
        ws['B2'].value = 'Sagehen Creek'

    # Site updates

    if ws['B6'].value[:6] == 'COER12': # F12 --> Forest 12
        ws['B4'].value = 'Forest 12'

    if ws['B6'].value[:6] == 'COER13':  # F13 --> Forest 13
        ws['B4'].value = 'Forest 13'

    if ws['B6'].value[:6] == 'COER14':  # F14 --> Forest 14
        ws['B4'].value = 'Forest 14'

    if ws['B6'].value[:6] == 'COERO2':  # O2 --> Open 2
        ws['B4'].value = 'Open 2'

    if ws['B6'].value[:6] == 'COERO4': # O4 --> Open 4
        ws['B4'].value = 'Open 4'

    if ws['B6'].value[:6] == 'COERO6': # O6 --> Open 6
        ws['B4'].value = 'Open 6'

    if ws['B6'].value[:6] == 'COFEJ1':  # JPL#1 --> JPL 1
        ws['B4'].value = 'JPL 1'

    if ws['B6'].value[:6] == 'COFEJ2': # JPL#2 --> JPL 2
        ws['B4'].value = 'JPL 2'

    if ws['B6'].value[:6] == 'COFEB1': # SNB#1 --> SNB 1
        ws['B4'].value = 'SNB 1'

    if ws['B6'].value[:6] == 'COFEB2': # SNB#2 --> SNB 2
        ws['B4'].value = 'SNB 2'


    # Save workbook
    wb.save(filename)

print('script is complete')
