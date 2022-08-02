import pandas as pd
import glob
import os
from datetime import datetime, timedelta
from csv import writer
from pathlib import Path
from openpyxl import load_workbook

'''
This script is to convert all fieldwork times to Local, Standard Time

Daylight Savings Time occurred on Sunday, March 8th, 2020, therefore all snow pits after
that date and converted to Standard Time (-1 hr from the fieldbook sheet)


e.g. Time recorded in the field was 11:24, the converted Standard Time would be 10:24
'''

path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/COMPLETE/')

daylight_savings = datetime(2020, 3, 8) # March 8th, 2020


for filename in sorted(path.rglob('*.xlsx')):

    # load workbook
    wb = load_workbook(filename)
    ws = wb.active

    # date and time variables
    date = ws['S6'].value
    time = ws['X6'].value # could be string or datetime.time type

    # convert string times to datetime types
    if isinstance(time, str):
        time = datetime.strptime(time, '%H:%M').time()

    dt = datetime.combine(date.date(), time)

    # convert to Local, Standard time if date is past daylight savings
    if date > daylight_savings:
        # convert to Local Standard Time
        newTime = dt - timedelta(hours=1) # this has to be a datetime combo, not just time

        # file management
        Time_str = filename.stem.split('_')[2] # old time string from filename
        newTime_str = newTime.time().strftime("%H%M") # new time as string from line above
        newFilename = Path(str(filename).replace(Time_str, newTime_str)) # update filename

        # overwrite time
        ws['X6'].value = newTime.time() # make it just time
        os.remove(filename) # remove original file ("replace" doesn't overwrite file)

        print("%s @ %s --> %s" %(date.date(), time, newTime.time())) #great print statement (notice all dates post 3/8 are -1-hr.)

    else: # technically this isn't necessary, but good way to make sure all times are datetime type (not strings)
        newTime = time #same variable to keep it easy, not all times are being changed!
        newFilename = filename # no change

        print("%s @ %s --> %s" %(date.date(), time, time)) # see, time doesn't change here

        # overwrite time (for all, why not)
        ws['X6'].value = newTime # it is datetime.time() already 

    # Save workbook
    wb.save(newFilename)

print('all done, did it work?')





# ~~~~~~~~~~~~~~~~~~~~might need later:
#if:
# pitID, date_str, time_str, desc = filename.stem.split('_') # split filename to isolate time_str
# newFilename = pitID + '_' + date_str + '_' + newTime.time().strftime("%H%M") + '_' + desc + filename.suffix # remove old time string, and replace with newTime
# newFilename = Path(pitID + '_' + date_str + '_' + newTime.time().strftime("%H%M") + '_' + desc + filename.suffix) # remove old time string, and replace with newTime
#
#else:
# newFilename = filename
