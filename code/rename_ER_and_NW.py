import datetime
import glob
import os
import shutil
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook


location = 'Niwot' # East-River
path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/{}-formatted'.format(location))

for filename in path.rglob('*.xlsx'):

    # extract pieces from original filename (provided by Hannah B.)
    directory = filename.parent #dir
    ext = filename.suffix #ext
    name = filename.stem #stem
    pitID, dateString, desc = [name.split('_')[i] for i in (0,1,5)]

    # get time from the .xlsx file (pit sheet file)
    wb = load_workbook(filename)
    ws = wb.active
    time = ws['X6'].value # some are strings, some are datetime....

    if isinstance(time, str): # if string, pad with 0 in front (e.g. (9:45 --> 09:45))
        if len(time) < 5:
            time = time.zfill(5)

    if isinstance(time, datetime.time): # convert datetime.time --> string
        time = datetime.time.strftime(time, "%H%M")

    # create "new" file name to match SnowEx form
    newname = pitID + '_' +  dateString + '_' + time.replace(":","") + '_' + desc + ext
    print("%s --> %s" %(name, newname))
    filename.rename(Path(directory, newname))
