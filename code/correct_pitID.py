import pandas as pd
import glob
from csv import writer
from pathlib import Path
from openpyxl import load_workbook



path = Path('/Users/meganmason491/Google Drive/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/Fraser-EDITED/')


for filename in path.rglob('*.xlsx'):
    print(filename.stem)
    new_pitID = filename.name[:15] # correct pitID, and all of date

    wb = load_workbook(filename)
    ws = wb.active #grave active worksheet
    ws['B6:F6'][0][0].value = new_pitID
    wb.save(filename)
