import pandas as pd
import glob
from csv import writer
from pathlib import Path
from openpyxl import load_workbook



path = Path('/Users/meganmason491/Google Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/Upper-Gunnison-EDITED')


for filename in path.rglob('*.xlsx'):
    print(filename.stem)
    new_location = 'East River'
    new_pitID = filename.name[:15] # correct pitID, and all of date

    wb = load_workbook(filename)
    ws = wb.active #grave active worksheet
    ws['B2:K2'][0][0].value = new_location
    ws['B6:F6'][0][0].value = new_pitID
    wb.save(filename)
