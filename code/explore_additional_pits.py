import datetime
import glob
import os
import shutil
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

er_pits = Path('/Users/meganmason491/Google Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_fromPIs/East-River/WY 2020')
nw_pits = Path('/Users/meganmason491/Google Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_fromPIs/Niwot/WY 2020')

def wordfinder(filename, searchString):

    wb = load_workbook(filename)
    ws = wb.active

    if wb.active.title == "ENVIRONMENT": #strange one pit (so far) has the "environment" sheet as active, vers. the "PIT" sheet in East River pits
        ws = wb['PIT']


    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if searchString == ws.cell(i,j).value:
                cell = ws.cell(i,j) # "cell object"
                cell = str(cell) # convert to string, maybe something better here?
                cell = cell.split('.')[-1] # save the last part
                cell = cell[:2] # grab column, row

                char_list = [char for char in cell] # split characters
                cell = ''.join(char_list[0] + str(int(char_list[1]) + 1)) # add 1 to the row, and rejoin (e.g. cell=B1 -> B2)

                return cell
                # print(ws.cell(i,j))

def readPits(filename, l, s, p, o, d, t):
    #load
    wb = load_workbook(filename)
    ws = wb.active

    #location
    location = ws[l].value
    # if location == None:
    #     location = ''

    # site
    site = ws[s].value
    # if site == None:
    #     site = ''

    # pitID
    pitID = ws[p].value
    # if pitID == None:
    #     pitID = ''

    #surveyors
    surveyor = ws[o].value
    # if surveyor == None:
    #     surveyor = ''

    # date
    dateStr = ws[d].value
    if dateStr == None:
        date = ''
    else:
        date_parse = datetime.datetime.strptime(dateStr, '%m/%d/%Y')
        date = datetime.datetime.strftime(date_parse, '%Y-%m-%d')

    # time
    time = ws[t].value
    # if timeStr == None:
    #     time = ''
    # else:
    #     time = datetime.datetime.strftime(timeStr, '%H:%M')


    return filename.stem, pitID, location, site, surveyor, date, time


data = []
column_lst = ['Filename', 'PitID', 'Location', 'Site', 'Surveyor', 'Date', 'Time']


for filename in er_pits.rglob('*.xlsx'): #20200201_CB_O6_2
    print(filename.stem)

    l = wordfinder(filename, "Location:")
    s = wordfinder(filename, "Site:")
    p = wordfinder(filename, "Pit ID:")
    o = wordfinder(filename, "Surveyors:")
    d = wordfinder(filename, "Date:")
    t = wordfinder(filename, "Time:")

    data.append(readPits(filename, l, s, p, o, d, t))

# convert data list to dataframe
df = pd.DataFrame(data, columns=column_lst)

print('No. of data rows:', len(df.index))

# create csv
df.to_csv('/Users/meganmason491/Documents/snowex/2020/timeseries/qaqc_pits/additional_pit_stats.csv', sep=',', header=True)
