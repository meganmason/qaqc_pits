import pandas as pd
import glob
from datetime import datetime
from csv import writer
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font

'''
This script is to swap out "surveyor" for "observer" to retroactivily match the 2021 snow pits.
'''

# path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/COMPLETE')
path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/interval-boards-only-sheets')


for filename in path.rglob('*.xlsx'):

    # load workbook (excel pit sheet)
    wb = load_workbook(filename)
    ws = wb.active

    # swap "Surveyors" for "Observers"
    ws['L1'].value = "Observers:"
    ws['L1'].font = Font(bold=True)

# Save workbook
    wb.save(filename)

print('all done')
