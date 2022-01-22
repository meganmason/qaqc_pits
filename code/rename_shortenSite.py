import pandas as pd
import glob
from csv import writer
from pathlib import Path
from openpyxl import load_workbook


# knobs
which_folder = 'Little-Cottonwood-EDITED/atwater'
site_name_update = 'Atwater'
path = Path('/Users/meganmason491/Google Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/{}'.format(which_folder))


for filename in path.rglob('*.xlsx'):
    print(filename.stem)
    new_site = site_name_update


    wb = load_workbook(filename)
    ws = wb.active #grave active worksheet
    ws['B4:K4'][0][0].value = new_site
    wb.save(filename)
