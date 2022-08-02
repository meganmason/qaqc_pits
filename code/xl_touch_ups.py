import pandas as pd
import glob
from csv import writer
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# path = Path('/Users/meganmason491/Documents/snowex/2020/timeseries/qaqc_pits/pits_csv_edited/pits/')
path = Path('/Users/mamason6/Documents/snowex/core-datasets/ground/snow-pits/outputs/pits/')

for filename in path.rglob('*.xlsx'):
    print(filename.name)

# Grain Size column spacing
    wb = load_workbook(filename)
    ws = wb.active #grave active worksheet
    ws.column_dimensions['R'].width = 4.33 #(original for the 2-4 mm width)
    ws.column_dimensions['S'].width = 4.33
    ws.column_dimensions['T'].width = 4.33
    ws.column_dimensions['U'].width = 4.33
    ws.column_dimensions['V'].width = 4.33

# Dielectric constant alignment
    ws['H9'].alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
    ws['J9'].alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)

# Shrink to fit
    # ws['L8'].alignment = Alignment(shrink_to_fit=True)

# Save workbook
    wb.save(filename)
