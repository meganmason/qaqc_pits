import pandas as pd
import glob
from datetime import datetime
from csv import writer
from pathlib import Path
from openpyxl import load_workbook

'''
This script is to append time to the pitID.

Uniquie pit ID has a:
    - 6-letter code
    - date
    - time

e.g. IDBRBU_20200126_1456
'''

path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/COMPLETE')
# path = Path('/Users/mamason6/Google Drive/My Drive/SnowEx-2020/SnowEx-2020-timeseries-pits/timeseries_pitbook_sheets_EDIT/FOR_EDIT/Grand-Mesa-2')

for filename in path.rglob('*.xlsx'):

    # extract time from filename components
    time = filename.stem.split('_')[2]

    # load workbook (excel pit sheet)
    wb = load_workbook(filename)
    ws = wb.active

    # append time to pitID
    pitID = ws['B6'].value
    newPitID = pitID[:15] + '_' + time
    ws['B6'].value = newPitID
    print("%s --> %s" %(pitID, newPitID))

# Save workbook
    wb.save(filename)
